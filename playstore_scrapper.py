# -*- coding: utf-8 -*-
"""
Created on Sun May 19 12:15:01 2019

@author: praty
"""
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import openpyxl as opxl

Top_100_Mobile = opxl.Workbook()

Android_Top_100_free_us = Top_100_Mobile.get_sheet_by_name('Sheet')
Android_Top_100_free_us.title = "Android Top 100 Free US"

Android_Top_100_paid_us = Top_100_Mobile.create_sheet('Android Top 100 Paid US')
Android_Top_100_free_jp = Top_100_Mobile.create_sheet('Android Top 100 Free JP')
Android_Top_100_paid_jp = Top_100_Mobile.create_sheet('Android Top 100 Paid JP')




url_List = ['https://play.google.com/store/apps/category/GAME/collection/topselling_free?gl=us', 'https://play.google.com/store/apps/category/GAME/collection/topselling_paid?gl=us', 'https://play.google.com/store/apps/category/GAME/collection/topselling_free?gl=jp', 'https://play.google.com/store/apps/category/GAME/collection/topselling_paid?gl=jp']
List = []


for url_Number in range(4):
    browser = webdriver.Firefox()
    browser.get(url_List[url_Number])
    
    elem = browser.find_element_by_tag_name('html')
    elem.send_keys(Keys.END)
    time.sleep(5)
    elem.send_keys(Keys.END)
    
    
    """
    database=browser.find_element_by_css_selector('#body-content > div:nth-child(1) > div > div.main-content > div > div > div > div.id-card-list.card-list.two-cards > div:nth-child(11) > div > div.details > a.title')
    database.get_attribute('title')
    database.get_attribute('href')
    """
    List = []
    
    for game_rank in range(100):
        database=browser.find_element_by_css_selector('#body-content > div:nth-child(1) > div > div.main-content > div > div > div > div.id-card-list.card-list.two-cards > div:nth-child('+str(game_rank+1)+') > div > div.details > a.title')
        List.append([database.get_attribute('title')])
        List[game_rank].append(database.get_attribute('href'))
    
    if url_Number == 0:
        Android_Top_100_free_us['A1'] = "Game Name"
        Android_Top_100_free_us['B1'] = "URL"
        for i in range(100):
            Android_Top_100_free_us['A' + str(i+2)] = List[i][0]
            Android_Top_100_free_us['B' + str(i+2)] = List[i][1]
    
    if url_Number == 1:
        Android_Top_100_paid_us['A1'] = "Game Name"
        Android_Top_100_paid_us['B1'] = "URL"
        for i in range(100):
            Android_Top_100_paid_us['A' + str(i+2)] = List[i][0]
            Android_Top_100_paid_us['B' + str(i+2)] = List[i][1]
    
    if url_Number == 2:
        Android_Top_100_free_jp['A1'] = "Game Name"
        Android_Top_100_free_jp['B1'] = "URL"
        for i in range(100):
            Android_Top_100_free_jp['A' + str(i+2)] = List[i][0]
            Android_Top_100_free_jp['B' + str(i+2)] = List[i][1]
    
    if url_Number == 3:
        Android_Top_100_paid_jp['A1'] = "Game Name"
        Android_Top_100_paid_jp['B1'] = "URL"
        for i in range(100):
            Android_Top_100_paid_jp['A' + str(i+2)] = List[i][0]
            Android_Top_100_paid_jp['B' + str(i+2)] = List[i][1]


Top_100_Mobile.save('F:\\academic\\jobs\\jobs\\gameopedia\\work\\Automation\\Scrapper\\Games_List.xlsx')

